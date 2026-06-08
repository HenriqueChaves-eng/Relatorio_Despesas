from __future__ import annotations

import os
from datetime import date
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


DATE_COLUMNS = ("D", "F", "H", "J", "L", "N", "P", "R")
CATEGORY_ROWS = {
    "Hospedagem": 21,
    "Refeicoes": 24,
    "Refeições": 24,
    "Pedagio": 27,
    "Pedágio": 27,
    "Frete": 30,
    "Combustivel": 33,
    "Combustível": 33,
    "Material de Uso e Consumo": 36,
    "Locacoes de Veiculos": 39,
    "Locações de Veículos": 39,
    "Transporte/taxi": 42,
    "Transporte/ taxi": 42,
    "Outras Despesas": 45,
}


def _anchor_cell(worksheet: Any, address: str) -> Any:
    cell = worksheet[address]
    if not isinstance(cell, MergedCell):
        return cell

    for merged_range in worksheet.merged_cells.ranges:
        if address in merged_range:
            return worksheet.cell(merged_range.min_row, merged_range.min_col)
    return cell


def _clear(worksheet: Any, address: str) -> None:
    _anchor_cell(worksheet, address).value = None


def _set_text(worksheet: Any, address: str, value: object) -> None:
    text = "" if value is None else str(value).strip()
    _anchor_cell(worksheet, address).value = text or None


def _set_number(worksheet: Any, address: str, value: object) -> None:
    _anchor_cell(worksheet, address).value = None if value in (None, "") else float(value)


def _set_date(worksheet: Any, address: str, value: object) -> None:
    cell = _anchor_cell(worksheet, address)
    if value in (None, ""):
        cell.value = None
        return

    cell.value = date.fromisoformat(str(value))
    cell.number_format = "dd/mm/yyyy"


def _set_formula_or_number(worksheet: Any, address: str, value: object, formula: object) -> None:
    formula_text = "" if formula is None else str(formula).strip()
    if formula_text:
        _anchor_cell(worksheet, address).value = formula_text if formula_text.startswith("=") else f"={formula_text}"
        return
    _set_number(worksheet, address, value)


def generate_excel_report(payload: dict[str, object]) -> Path:
    template_path = Path(str(payload["template_path"]))
    output_dir = Path(str(payload["output_dir"]))
    xlsx_path = Path(str(payload["xlsx_path"]))
    dates = list(payload.get("dates") or [])
    expense_cells = list(payload.get("expense_cells") or [])
    other_expenses = list(payload.get("other_expenses") or [])
    mileage_items = list(payload.get("mileage") or [])

    if not template_path.exists():
        raise FileNotFoundError(f"Modelo não encontrado: {template_path}")
    if len(dates) > 8:
        raise ValueError("O modelo aceita no máximo 8 datas de despesas.")
    if len(other_expenses) > 4:
        raise ValueError("O modelo aceita no máximo 4 linhas em Outras Despesas.")
    if len(mileage_items) > 4:
        raise ValueError("O modelo aceita no máximo 4 linhas em Despesas de Quilometragem.")

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template_path)
    worksheet = workbook.worksheets[0]

    for column in DATE_COLUMNS:
        _clear(worksheet, f"{column}19")
        for row in (21, 24, 27, 30, 33, 36, 39, 42, 45):
            _clear(worksheet, f"{column}{row}")

    for row in (51, 52, 53, 54):
        _clear(worksheet, f"B{row}")
        _clear(worksheet, f"D{row}")
        _clear(worksheet, f"X{row}")

    for row in (58, 59, 60, 61):
        _clear(worksheet, f"B{row}")
        _clear(worksheet, f"D{row}")
        _clear(worksheet, f"G{row}")

    for address in ("K65", "C67", "O65", "S65"):
        _clear(worksheet, address)

    employee = dict(payload.get("employee") or {})
    trip = dict(payload.get("trip") or {})
    _set_text(worksheet, "B8", employee.get("name"))
    _set_text(worksheet, "I8", employee.get("cpf"))
    _set_text(worksheet, "B10", employee.get("cost_center"))
    _set_text(worksheet, "B12", employee.get("bank"))
    _set_text(worksheet, "K12", employee.get("agency"))
    _set_text(worksheet, "P12", employee.get("account"))
    _set_date(worksheet, "E10", trip.get("start_date"))
    _set_date(worksheet, "I10", trip.get("end_date"))
    _set_text(worksheet, "B14", trip.get("reason"))
    _set_date(worksheet, "C67", trip.get("report_date"))

    for index, expense_date in enumerate(dates):
        _set_date(worksheet, f"{DATE_COLUMNS[index]}19", expense_date)

    for entry_value in expense_cells:
        entry = dict(entry_value)
        category = str(entry.get("category") or "")
        if category not in CATEGORY_ROWS:
            raise ValueError(f"Categoria não mapeada: {category}")
        _set_formula_or_number(
            worksheet,
            f"{entry.get('column')}{CATEGORY_ROWS[category]}",
            entry.get("value"),
            entry.get("formula"),
        )

    for row in (21, 24, 27, 30, 33, 36, 39, 42):
        worksheet[f"T{row}"] = f'=IF(SUM(D{row}:S{row + 2})=0,"",SUM(D{row}:S{row + 2}))'
    worksheet["T45"] = '=IF(SUM(D45:R45)=0,"",SUM(D45:R45))'

    for first, last in (("D", "E"), ("F", "G"), ("H", "I"), ("J", "K"), ("L", "M"), ("N", "O"), ("P", "Q"), ("R", "S")):
        worksheet[f"{first}47"] = f"=SUM({first}21:{last}46)"
    worksheet["T47"] = "=SUM(D47:S48)"

    for row, item_value in zip((51, 52, 53, 54), other_expenses):
        item = dict(item_value)
        _set_date(worksheet, f"B{row}", item.get("date"))
        _set_text(worksheet, f"D{row}", item.get("description"))
        _set_number(worksheet, f"X{row}", item.get("value"))

    for row, item_value in zip((58, 59, 60, 61), mileage_items):
        item = dict(item_value)
        _set_date(worksheet, f"B{row}", item.get("date"))
        _set_number(worksheet, f"D{row}", item.get("distance"))
        _set_text(worksheet, f"G{row}", item.get("description"))

    _set_number(worksheet, "S61", payload.get("km_rate"))
    for row in (58, 59, 60, 61):
        worksheet[f"X{row}"] = f"=D{row}*$S$61"
    worksheet["X62"] = "=SUM(X58:Y61)"

    _set_number(worksheet, "K65", payload.get("advance"))
    worksheet["O65"] = "=T47+X62"
    worksheet["S65"] = "=O65-K65"

    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"

    temp_path = xlsx_path.with_name(f".{xlsx_path.name}.{uuid4().hex}.tmp.xlsx")
    try:
        workbook.save(temp_path)
        workbook.close()
        validation = load_workbook(temp_path, read_only=True, data_only=False)
        try:
            if not validation.sheetnames:
                raise ValueError("O Excel gerado não possui planilhas.")
        finally:
            validation.close()
        os.replace(temp_path, xlsx_path)
    finally:
        workbook.close()
        temp_path.unlink(missing_ok=True)
    return xlsx_path
